# core_scraper_xls.py
import io
from typing import Dict, List, Tuple
from openpyxl import load_workbook
import xlrd  # 1.2.0
from bs4 import BeautifulSoup

from utils import parse_price, to_int_loose, txt
from db import (
    ensure_product_and_alias,
    upsert_stock,
    begin_ingestion_run,
    insert_inventory_raw,
    bulk_insert_inventory_raw,
    bulk_upsert_inventory_current,
)

# -------------------- Lectores --------------------

def _normalize_headers(headers: List[str]) -> List[str]:
    # Normaliza y pasa a minúsculas para que no importe la capitalización
    return [str(h or "").strip().lower() for h in headers]

def _read_html_table(data: bytes) -> Tuple[List[str], List[List]]:
    """
    Lee la PRIMERA tabla de un HTML (muchos sistemas exportan HTML con extensión .xls).
    """
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1", errors="ignore")

    soup = BeautifulSoup(text, "lxml")
    table = soup.find("table")
    if not table:
        raise ValueError("No se encontró <table> en el HTML exportado.")

    trs = table.find_all("tr")
    rows = []
    for tr in trs:
        cells = [c.get_text(strip=True) for c in (tr.find_all("td") or tr.find_all("th"))]
        if cells:
            rows.append(cells)

    if not rows:
        return [], []

    headers = _normalize_headers(rows[0])
    body = [r for r in rows[1:] if any(x.strip() for x in r)]
    return headers, body

def _read_xlsx(data: bytes) -> Tuple[List[str], List[List]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[(c if c is not None else "") for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return [], []
    headers = _normalize_headers(rows[0])
    body = [r for r in rows[1:] if any(str(x).strip() for x in r)]
    return headers, body

def _read_xls(data: bytes) -> Tuple[List[str], List[List]]:
    # Muchos “.xls” de sistemas realmente son HTML; detecta eso primero
    head = data[:256].lower()
    if b"<html" in head or b"<table" in head:
        return _read_html_table(data)

    book = xlrd.open_workbook(file_contents=data)
    sh = book.sheet_by_index(0)
    rows = []
    for rx in range(sh.nrows):
        row = []
        for cx in range(sh.ncols):
            v = sh.cell_value(rx, cx)
            row.append(v)
        rows.append(row)

    if not rows:
        return [], []
    headers = _normalize_headers(rows[0])
    body = [r for r in rows[1:] if any(str(x).strip() for x in r)]
    return headers, body

def _detect_and_read(data: bytes, name: str) -> Tuple[List[str], List[List]]:
    low = data[:4096].lower()
    if b"<html" in low or b"<table" in low or b"<form" in low:
        return _read_html_table(data)

    if name.endswith(".xlsx") or data[:2] == b"PK":
        return _read_xlsx(data)
    # por defecto intenta xls (y dentro detecta html camuflado)
    return _read_xls(data)

# -------------------- Proceso e inserción --------------------

def process_spreadsheet(data: bytes, suggested_name: str, system_code: str, store_id: int, colmap: Dict[str, str]) -> int:
    headers, rows = _detect_and_read(data, suggested_name)
    idx = {h: i for i, h in enumerate(headers)}

    # Normaliza el mapeo a minúsculas (coincide con headers ya normalizados)
    colmap = {k: str(v).strip().lower() for k, v in colmap.items()}

    # valida columnas requeridas
    for k, colname in colmap.items():
        if colname not in idx:
            raise RuntimeError(f"Columna '{colname}' no encontrada en el archivo. Encabezados: {headers}")

    run_id = begin_ingestion_run(system_code, store_id)
    total = 0

    raw_rows = []
    current_rows = []

    for i, r in enumerate(rows, start=1):
        system_id = txt(r[idx[colmap["system_id"]]])
        sku       = txt(r[idx[colmap["sku"]]])
        name      = txt(r[idx[colmap["name"]]])
        existencia= to_int_loose(r[idx[colmap["existencia"]]])
        costo     = parse_price(r[idx[colmap["costo"]]])
        precio    = parse_price(r[idx[colmap["precio"]]])

        # Acumula crudo (incluye duplicados)
        raw_rows.append(dict(
            system_code=system_code,
            store_id=store_id,
            system_id=system_id or f"row-{i}",
            sku=sku,
            name=name,
            existencia=existencia,
            costo=costo,
            precio=precio,
        ))

        if system_id:
            current_rows.append(dict(
                system_code=system_code,
                store_id=store_id,
                system_id=system_id,
                sku=sku,
                name=name,
                existencia=existencia,
                costo=costo,
                precio=precio,
                seen_at=None,
            ))
            total += 1

    # Inserciones en bloque para minimizar overhead de conexiones/roundtrips
    if raw_rows:
        bulk_insert_inventory_raw(raw_rows)
    if current_rows:
        bulk_upsert_inventory_current(current_rows)

    return total
